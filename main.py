import time
from datetime import datetime
import pandas as pd
import streamlit as st
import gspread
from google.oauth2.service_account import Credentials
from gspread.utils import rowcol_to_a1
import numpy as np
import io
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ──────────────────────────────────────────────
# Connection & UI Layer
# ──────────────────────────────────────────────

def get_gsheet_conn():
    creds_dict = {
        "type": st.secrets["connections"]["gsheets"]["type"],
        "project_id": st.secrets["connections"]["gsheets"]["project_id"],
        "private_key_id": st.secrets["connections"]["gsheets"]["private_key_id"],
        "private_key": st.secrets["connections"]["gsheets"]["private_key"],
        "client_email": st.secrets["connections"]["gsheets"]["client_email"],
        "client_id": st.secrets["connections"]["gsheets"]["client_id"],
        "auth_uri": st.secrets["connections"]["gsheets"]["auth_uri"],
        "token_uri": st.secrets["connections"]["gsheets"]["token_uri"],
    }

    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive"
    ]

    creds = Credentials.from_service_account_info(creds_dict, scopes=scopes)
    client = gspread.authorize(creds)
    return client


def connect_gsheet():
    try:
        client = get_gsheet_conn()
        SPREADSHEET_ID = "1MSKQUSdBTOEI2-YNE7Z43_SfbFsZxrOkD2UIBByX_bs"
        spreadsheet = client.open_by_key(SPREADSHEET_ID)
        print("Connection successful...!!!")
        return spreadsheet
    except Exception as e:
        print(f"Unable to connect google sheet: {e}")
        show_popup(f"Unable to connect google sheet: {e}", type="error")
        raise e


def show_popup(message, type="success"):
    if type == "success":
        st.toast(f"✅ {message}")
    elif type == "error":
        st.toast(f"❌ {message}")
    elif type == "warning":
        st.toast(f"⚠️ {message}")
    elif type == "info":
        st.toast(f"ℹ️ {message}")


# ──────────────────────────────────────────────
# Core Transformation (func1)
# ──────────────────────────────────────────────

def func1(raw_file):
    try:
        spreadsheet = connect_gsheet()

        try:
            detailData_worksheet = spreadsheet.worksheet("Detailed_Data")
        except gspread.WorksheetNotFound:
            detailData_worksheet = spreadsheet.add_worksheet("Detailed_Data", rows=5000, cols=30)
            
        detailData_worksheet.clear()

        data = pd.read_excel(raw_file)
        data.columns = data.columns.str.lower().str.replace(" ", "_").str.replace(".", "_").str.strip()
        
        # # Define the list of status codes you want to exclude
        # excluded_statuses = ["TO_BE_REJECTED", "RAN_C_CN_DUE", "RAN_D_CN_DUE"]

        # # Filter the dataframe to keep only rows NOT in that list
        # data = data[~data["status_code"].str.upper().str.strip().isin(excluded_statuses)]

        selected_columns = [
            "register_id", "job_id", "service_id", "customer_name",  "phone1","customer_type", "address1", "city", "state", "customer_pincode", "producttype_code", "model_code", "product_srno", "product_srno2", "company_name","provider_phone1", "circle", "customer_type", "call_date", "status_updated_date", "status_code",  "registration_date", "warrantytype", "invoice_no", "invoice_date"
        ]
        # Eliminate duplicate selected columns if any exist in array config
        selected_columns = list(dict.fromkeys(selected_columns))
        data = data[selected_columns].copy()
        
        data["service_id"] = data["service_id"].astype(str).str.strip()
        data["call_date"] = pd.to_datetime(data["call_date"]).dt.normalize()
        data["status_updated_date"] = pd.to_datetime(data["status_updated_date"]).dt.normalize()

        # Fix type safety: ensure comparison uses consistent date formats
        today_date = pd.Timestamp.now().normalize()
        data["today_date"] = today_date
        data["age_from_call_reg"] = data["today_date"] - data["call_date"]

        norms_worksheet = spreadsheet.worksheet("Norms_Data")
        norms_data = norms_worksheet.get_all_records()
        status_data = pd.DataFrame(norms_data)
        status_data.columns = status_data.columns.str.lower().str.strip().str.replace(" ", "_")

        merged_data = data.merge(status_data[["status", "team", "number"]], left_on="status_code", right_on="status", how="left")

        merged_data["age_reg_days"] = merged_data["age_from_call_reg"].dt.days.fillna(0).astype(int)
        
        # Correct: create boolean flag columns, cast True/False to 0/1
        merged_data["7+_calls"]  = ((merged_data["age_reg_days"] > 7)  & (merged_data["age_reg_days"] <= 14)).astype(int)

        merged_data["15+_calls"] = (merged_data["age_reg_days"] > 14).astype(int)
        
        if not merged_data.empty:
            data_to_write = [merged_data.columns.tolist()] + merged_data.fillna("").astype(str).values.tolist()
            detailData_worksheet.update(data_to_write)
            show_popup("Data stored in the database", type="success")
        else:
            show_popup("No data found after filtering...!", type="info")
            
        return merged_data

    except Exception as e:
        print(f"Error in func1: {e}")
        show_popup(f"Error in function is: {e}", type="error")


# ──────────────────────────────────────────────
# Vectorized Sync Engines
# ──────────────────────────────────────────────

def update_ch_raw_data():
    try:
        print("Inside update ch raw data function...")
        spreadsheet = connect_gsheet()

        detail_ws   = spreadsheet.worksheet("Detailed_Data")
        detail_data = detail_ws.get_all_values()
        if len(detail_data) <= 1:
            show_popup("Detailed_Data sheet is empty!", type="info")
            return

        df_detail = pd.DataFrame(detail_data[1:], columns=detail_data[0])
        df_detail.columns = df_detail.columns.str.lower().str.strip().str.replace(" ", "_")
        df_detail = df_detail.loc[:, ~df_detail.columns.duplicated()]
        df_detail["service_id"] = df_detail["service_id"].astype(str).str.strip()

        df_filtered = df_detail[df_detail["7+_calls"].astype(str) == "1"].copy()
        print("*****The dimension of filtered CH dataframe is:", df_filtered.shape)
        df_filtered = df_filtered.reset_index(drop=True)

        if df_filtered.empty:
            show_popup("No 7+ calls data found!", type="info")
            # return

        all_detail_ids    = set(df_detail["service_id"])
        ageing_sheet_name = "CH Raw Data"

        try:
            ageing_ws   = spreadsheet.worksheet(ageing_sheet_name)
            ageing_data = ageing_ws.get_all_values()

            if ageing_data and len(ageing_data) > 1:
                df_ageing_existing = pd.DataFrame(ageing_data[1:], columns=ageing_data[0])
                df_ageing_existing.columns = df_ageing_existing.columns.str.lower().str.strip().str.replace(" ", "_")
                df_ageing_existing = df_ageing_existing.loc[:, ~df_ageing_existing.columns.duplicated()].reset_index(drop=True)
                df_ageing_existing["service_id"] = df_ageing_existing["service_id"].astype(str).str.strip()
            else:
                df_ageing_existing = pd.DataFrame()
        except gspread.WorksheetNotFound:
            ageing_ws = spreadsheet.add_worksheet(ageing_sheet_name, rows=5000, cols=30)
            df_ageing_existing = pd.DataFrame()

        if df_ageing_existing.empty:
            data_to_write = [df_filtered.columns.tolist()] + df_filtered.fillna("").astype(str).values.tolist()
            ageing_ws.update(data_to_write)
            show_popup(f"CH Raw Data sheet created with {len(df_filtered)} records!", type="success")
            return

        # ── Move rows no longer in Detailed_Data → cc_ch_data ──
        existing_ageing_ids = set(df_ageing_existing["service_id"])
        moved_ids           = existing_ageing_ids - all_detail_ids
        
        # Preserves the intact records (including remarks) before clearing them out
        df_moved            = df_ageing_existing[df_ageing_existing["service_id"].isin(moved_ids)].copy()
        print("Rows to move → cc_ch_data:", df_moved.shape)

        # Deletes the old records entirely from the main CH raw dataframe state
        df_ageing_existing = df_ageing_existing[
            ~df_ageing_existing["service_id"].isin(moved_ids)
        ].reset_index(drop=True).copy()

        if not df_moved.empty:
            cc_sheet_name = "cc_ch_data"
            try:
                cc_ws   = spreadsheet.worksheet(cc_sheet_name)
                cc_data = cc_ws.get_all_values()
                if cc_data and len(cc_data) > 1:
                    df_cc_existing = pd.DataFrame(cc_data[1:], columns=cc_data[0])
                    df_cc_existing.columns = df_cc_existing.columns.str.lower().str.strip().str.replace(" ", "_")
                    df_cc_existing = df_cc_existing.loc[:, ~df_cc_existing.columns.duplicated()].reset_index(drop=True)
                    df_cc_existing["service_id"] = df_cc_existing["service_id"].astype(str).str.strip()
                else:
                    df_cc_existing = pd.DataFrame()
            except gspread.WorksheetNotFound:
                cc_ws = spreadsheet.add_worksheet(cc_sheet_name, rows=5000, cols=30)
                df_cc_existing = pd.DataFrame()

            if df_cc_existing.empty:
                # Direct bulk write when target sheet is empty
                cc_data_to_write = (
                    [df_moved.columns.tolist()]
                    + df_moved.fillna("").astype(str).values.tolist()
                )
                cc_ws.update(cc_data_to_write)
                print(f"cc_ch_data created with {len(df_moved)} rows.")
            else:
                # Append rows unique to this execution run
                existing_cc_ids = set(df_cc_existing["service_id"])
                df_moved_new    = df_moved[~df_moved["service_id"].isin(existing_cc_ids)].copy()

                if not df_moved_new.empty:
                    # FIX: Merging dataframes via concat to let unique columns (like remarks) flow in without deletion
                    df_cc_existing = pd.concat([df_cc_existing, df_moved_new], ignore_index=True)
                    df_cc_existing.fillna("", inplace=True)
                    
                    cc_final_data  = (
                        [df_cc_existing.columns.tolist()]
                        + df_cc_existing.astype(str).values.tolist()
                    )
                    cc_ws.clear()
                    cc_ws.update(cc_final_data)
                    print(f"cc_ch_data appended with {len(df_moved_new)} new rows (including remarks).")
                else:
                    print("All moved rows already exist in cc_ch_data — nothing appended.")

        # ── Vectorized Update Engine (merge-based, avoids index mismatch) ──
        ageing_cols  = df_ageing_existing.columns.tolist()
        remarks_cols = {col for col in ageing_cols if col.startswith("remark")}

        cols_to_update = [
            col for col in df_filtered.columns
            if col != "service_id"
            and col not in remarks_cols
            and col in df_ageing_existing.columns
        ]
        print(f"Columns being updated: {cols_to_update}")

        matched_ids   = set(df_ageing_existing["service_id"]) & set(df_filtered["service_id"])
        updated_count = len(matched_ids)
        df_ageing_new = df_filtered[~df_filtered["service_id"].isin(matched_ids)].reset_index(drop=True).copy()

        df_updates = (
            df_filtered[df_filtered["service_id"].isin(matched_ids)][["service_id"] + cols_to_update]
            .drop_duplicates(subset="service_id")
            .reset_index(drop=True)
        )

        df_ageing_existing = df_ageing_existing.merge(
            df_updates,
            on="service_id",
            how="left",
            suffixes=("", "_new")
        )

        for col in cols_to_update:
            new_col = col + "_new"
            if new_col in df_ageing_existing.columns:
                mask = df_ageing_existing[new_col].notna() & (df_ageing_existing[new_col] != "")
                df_ageing_existing.loc[mask, col] = df_ageing_existing.loc[mask, new_col]
                df_ageing_existing.drop(columns=[new_col], inplace=True)

        if not df_ageing_new.empty:
            df_ageing_new      = df_ageing_new.reindex(columns=ageing_cols, fill_value="")
            df_ageing_existing = pd.concat([df_ageing_existing, df_ageing_new], ignore_index=True)

        # Single bulk write to refresh main sheet state (safely removing the dropped records)
        final_data = [df_ageing_existing.columns.tolist()] + df_ageing_existing.fillna("").astype(str).values.tolist()
        ageing_ws.clear()
        ageing_ws.update(final_data)

        show_popup(
            f"CH Raw Data updated! {updated_count} rows updated, "
            f"{len(df_ageing_new)} new rows added, {len(df_moved)} rows moved to cc_ch_data.",
            type="success"
        )

    except Exception as e:
        print(f"Error in update_ch_raw_data: {e}")
        show_popup(f"Error While Updating CH Raw Data: {e}", type="error")


# def update_ho_raw_data():
#     try:
#         spreadsheet = connect_gsheet()
#         detail_ws = spreadsheet.worksheet("Detailed_Data")
#         detail_data = detail_ws.get_all_values()
#         if len(detail_data) <= 1:
#             show_popup("Detailed_Data sheet is empty!", type="info")
#             return

#         df_detail = pd.DataFrame(detail_data[1:], columns=detail_data[0])
#         df_detail.columns = df_detail.columns.str.lower().str.strip().str.replace(" ", "_")
#         df_detail = df_detail.loc[:, ~df_detail.columns.duplicated()]
#         df_detail["service_id"] = df_detail["service_id"].astype(str).str.strip()

#         df_filtered = df_detail[df_detail["15+_calls"].astype(str) == "1"].copy()
#         df_filtered  = df_filtered.reset_index(drop=True)
        
#         if df_filtered.empty:
#             show_popup("No 15+ calls data found!", type="info")
#             return

#         all_detail_ids = set(df_detail["service_id"])
#         ho_sheet_name  = "HO Raw Data"

#         try:
#             # To fetch the HO Data
#             ho_ws   = spreadsheet.worksheet(ho_sheet_name)
#             ho_data = ho_ws.get_all_values()

#             if ho_data and len(ho_data) > 1:
#                 df_ho_existing = pd.DataFrame(ho_data[1:], columns=ho_data[0])
#                 df_ho_existing.columns = df_ho_existing.columns.str.lower().str.strip().str.replace(" ", "_")
#                 df_ho_existing = df_ho_existing.loc[:, ~df_ho_existing.columns.duplicated()].reset_index(drop=True)
#                 df_ho_existing["service_id"] = df_ho_existing["service_id"].astype(str).str.strip()
#             else:
#                 df_ho_existing = pd.DataFrame()
#         except gspread.WorksheetNotFound:
#             ho_ws = spreadsheet.add_worksheet(ho_sheet_name, rows=5000, cols=30)
#             df_ho_existing = pd.DataFrame()

#         if df_ho_existing.empty:
#             data_to_write = [df_filtered.columns.tolist()] + df_filtered.fillna("").astype(str).values.tolist()
#             ho_ws.update(data_to_write)
#             show_popup(f"HO Raw Data sheet created with {len(df_filtered)} records!", type="success")
#             return

#         # ── Move rows no longer in Detailed_Data → cc_ho_data ──
#         existing_ho_ids = set(df_ho_existing["service_id"])
#         moved_ids       = existing_ho_ids - all_detail_ids
        
#         # This keeps the remarks intact from the main sheet data state
#         df_moved        = df_ho_existing[df_ho_existing["service_id"].isin(moved_ids)].copy()

#         # This cleanly drops the rows (including their remarks) from the live HO sheet state
#         df_ho_existing  = df_ho_existing[
#             ~df_ho_existing["service_id"].isin(moved_ids)
#         ].reset_index(drop=True).copy()

#         print(f"Rows to move → cc_ho_data: {len(df_moved)}")

#         if not df_moved.empty:
#             cc_sheet_name = "cc_ho_data"
#             try:
#                 cc_ws   = spreadsheet.worksheet(cc_sheet_name)
#                 cc_data = cc_ws.get_all_values()
#                 if cc_data and len(cc_data) > 1:
#                     df_cc_existing = pd.DataFrame(cc_data[1:], columns=cc_data[0])
#                     df_cc_existing.columns = df_cc_existing.columns.str.lower().str.strip().str.replace(" ", "_")
#                     df_cc_existing = df_cc_existing.loc[:, ~df_cc_existing.columns.duplicated()].reset_index(drop=True)
#                     df_cc_existing["service_id"] = df_cc_existing["service_id"].astype(str).str.strip()
#                 else:
#                     df_cc_existing = pd.DataFrame()
#             except gspread.WorksheetNotFound:
#                 cc_ws = spreadsheet.add_worksheet(cc_sheet_name, rows=5000, cols=30)
#                 df_cc_existing = pd.DataFrame()

#             if df_cc_existing.empty:
#                 cc_data_to_write = (
#                     [df_moved.columns.tolist()]
#                     + df_moved.fillna("").astype(str).values.tolist()
#                 )
#                 cc_ws.update(cc_data_to_write)
#                 print(f"cc_ho_data created with {len(df_moved)} rows.")
#             else:
#                 # Only extract records not already present in archive
#                 existing_cc_ids = set(df_cc_existing["service_id"])
#                 df_moved_new    = df_moved[~df_moved["service_id"].isin(existing_cc_ids)].copy()

#                 if not df_moved_new.empty:
#                     # FIX: Use dynamic concatenation instead of destructive .reindex()
#                     # This allows columns unique to df_moved (like remarks) to merge into the output schema safely
#                     df_cc_existing = pd.concat([df_cc_existing, df_moved_new], ignore_index=True)
#                     df_cc_existing.fillna("", inplace=True)
                    
#                     cc_final_data  = (
#                         [df_cc_existing.columns.tolist()]
#                         + df_cc_existing.astype(str).values.tolist()
#                     )
#                     cc_ws.clear()
#                     cc_ws.update(cc_final_data)
#                     print(f"cc_ho_data appended with {len(df_moved_new)} new rows (including remarks).")
#                 else:
#                     print("All moved rows already exist in cc_ho_data — nothing appended.")

#         # ── Vectorized Update Engine (merge-based, avoids index mismatch) ──
#         ho_cols      = df_ho_existing.columns.tolist()
#         remarks_cols = {col for col in ho_cols if col.startswith("remark")}

#         cols_to_update = [
#             col for col in df_filtered.columns
#             if col != "service_id"
#             and col not in remarks_cols
#             and col in df_ho_existing.columns
#         ]
#         print(f"Columns being updated: {cols_to_update}")

#         matched_ids   = set(df_ho_existing["service_id"]) & set(df_filtered["service_id"])
#         updated_count = len(matched_ids)
#         df_ho_new     = df_filtered[~df_filtered["service_id"].isin(matched_ids)].reset_index(drop=True).copy()

#         df_updates = (
#             df_filtered[df_filtered["service_id"].isin(matched_ids)][["service_id"] + cols_to_update]
#             .drop_duplicates(subset="service_id")
#             .reset_index(drop=True)
#         )

#         df_ho_existing = df_ho_existing.merge(
#             df_updates,
#             on="service_id",
#             how="left",
#             suffixes=("", "_new")
#         )

#         for col in cols_to_update:
#             new_col = col + "_new"
#             if new_col in df_ho_existing.columns:
#                 mask = df_ho_existing[new_col].notna() & (df_ho_existing[new_col] != "")
#                 df_ho_existing.loc[mask, col] = df_ho_existing.loc[mask, new_col]
#                 df_ho_existing.drop(columns=[new_col], inplace=True)

#         if not df_ho_new.empty:
#             df_ho_new  = df_ho_new.reindex(columns=ho_cols, fill_value="")
#             df_ho_existing = pd.concat([df_ho_existing, df_ho_new], ignore_index=True)

#         final_data = [df_ho_existing.columns.tolist()] + df_ho_existing.fillna("").astype(str).values.tolist()
#         ho_ws.clear()
#         ho_ws.update(final_data)

#         show_popup(
#             f"HO Raw Data updated! {updated_count} rows updated, "
#             f"{len(df_ho_new)} new rows added, {len(df_moved)} rows moved to cc_ho_data.",
#             type="success"
#         )

#     except Exception as e:
#         print(f"Error in update_ho_raw_data function: {e}")
#         show_popup(f"Error While Updating HO Data: {e}", type="error")



def update_ho_raw_data():
    try:
        spreadsheet = connect_gsheet()
        detail_ws = spreadsheet.worksheet("Detailed_Data")
        detail_data = detail_ws.get_all_values()
        
        if len(detail_data) <= 1:
            show_popup("Detailed_Data sheet is empty!", type="info")
            return

        df_detail = pd.DataFrame(detail_data[1:], columns=detail_data[0])
        df_detail.columns = df_detail.columns.str.lower().str.strip().str.replace(" ", "_")
        df_detail = df_detail.loc[:, ~df_detail.columns.duplicated()]
        df_detail["service_id"] = df_detail["service_id"].astype(str).str.strip()

        # ── NEW: Calculate and Insert Call Category Column ──
        if "age_reg_days" in df_detail.columns:
            # Safely convert to numeric for accurate condition checking
            age_numeric = pd.to_numeric(df_detail["age_reg_days"], errors="coerce")
            
            # Apply your logic: > 15 -> red call, == 15 -> encroaching
            df_detail["call_category"] = np.select(
                [age_numeric > 15, age_numeric == 15],
                ["red call", "encroaching"],
                default=""
            )
            
            # Reposition 'call_category' immediately after 'age_reg_days'
            idx = df_detail.columns.get_loc("age_reg_days") + 1
            cols = df_detail.columns.tolist()
            cols.insert(idx, cols.pop(cols.index("call_category")))
            df_detail = df_detail[cols]

        df_filtered = df_detail[df_detail["15+_calls"].astype(str) == "1"].copy()
        df_filtered = df_filtered.reset_index(drop=True)
        
        if df_filtered.empty:
            show_popup("No 15+ calls data found!", type="info")
            return

        all_detail_ids = set(df_detail["service_id"])
        ho_sheet_name = "HO Raw Data"

        try:
            # To fetch the HO Data
            ho_ws = spreadsheet.worksheet(ho_sheet_name)
            ho_data = ho_ws.get_all_values()

            if ho_data and len(ho_data) > 1:
                df_ho_existing = pd.DataFrame(ho_data[1:], columns=ho_data[0])
                df_ho_existing.columns = df_ho_existing.columns.str.lower().str.strip().str.replace(" ", "_")
                df_ho_existing = df_ho_existing.loc[:, ~df_ho_existing.columns.duplicated()].reset_index(drop=True)
                df_ho_existing["service_id"] = df_ho_existing["service_id"].astype(str).str.strip()
                
                # Ensure existing sheet gets the schema update if it's missing the column
                if "call_category" not in df_ho_existing.columns and "age_reg_days" in df_ho_existing.columns:
                    idx = df_ho_existing.columns.get_loc("age_reg_days") + 1
                    df_ho_existing.insert(idx, "call_category", "")
            else:
                df_ho_existing = pd.DataFrame()
        except gspread.WorksheetNotFound:
            ho_ws = spreadsheet.add_worksheet(ho_sheet_name, rows=5000, cols=30)
            df_ho_existing = pd.DataFrame()

        if df_ho_existing.empty:
            data_to_write = [df_filtered.columns.tolist()] + df_filtered.fillna("").astype(str).values.tolist()
            ho_ws.update(data_to_write)
            show_popup(f"HO Raw Data sheet created with {len(df_filtered)} records!", type="success")
            return

        # ── Move rows no longer in Detailed_Data → cc_ho_data ──
        existing_ho_ids = set(df_ho_existing["service_id"])
        moved_ids = existing_ho_ids - all_detail_ids
        
        # This keeps the remarks intact from the main sheet data state
        df_moved = df_ho_existing[df_ho_existing["service_id"].isin(moved_ids)].copy()

        # This cleanly drops the rows (including their remarks) from the live HO sheet state
        df_ho_existing = df_ho_existing[
            ~df_ho_existing["service_id"].isin(moved_ids)
        ].reset_index(drop=True).copy()

        print(f"Rows to move → cc_ho_data: {len(df_moved)}")

        if not df_moved.empty:
            cc_sheet_name = "cc_ho_data"
            try:
                cc_ws = spreadsheet.worksheet(cc_sheet_name)
                cc_data = cc_ws.get_all_values()
                if cc_data and len(cc_data) > 1:
                    df_cc_existing = pd.DataFrame(cc_data[1:], columns=cc_data[0])
                    df_cc_existing.columns = df_cc_existing.columns.str.lower().str.strip().str.replace(" ", "_")
                    df_cc_existing = df_cc_existing.loc[:, ~df_cc_existing.columns.duplicated()].reset_index(drop=True)
                    df_cc_existing["service_id"] = df_cc_existing["service_id"].astype(str).str.strip()
                    
                    # Ensure existing archive sheet gets the schema update if it's missing the column
                    if "call_category" not in df_cc_existing.columns and "age_reg_days" in df_cc_existing.columns:
                        idx = df_cc_existing.columns.get_loc("age_reg_days") + 1
                        df_cc_existing.insert(idx, "call_category", "")
                else:
                    df_cc_existing = pd.DataFrame()
            except gspread.WorksheetNotFound:
                cc_ws = spreadsheet.add_worksheet(cc_sheet_name, rows=5000, cols=30)
                df_cc_existing = pd.DataFrame()

            if df_cc_existing.empty:
                cc_data_to_write = (
                    [df_moved.columns.tolist()]
                    + df_moved.fillna("").astype(str).values.tolist()
                )
                cc_ws.update(cc_data_to_write)
                print(f"cc_ho_data created with {len(df_moved)} rows.")
            else:
                # Only extract records not already present in archive
                existing_cc_ids = set(df_cc_existing["service_id"])
                df_moved_new = df_moved[~df_moved["service_id"].isin(existing_cc_ids)].copy()

                if not df_moved_new.empty:
                    df_cc_existing = pd.concat([df_cc_existing, df_moved_new], ignore_index=True)
                    df_cc_existing.fillna("", inplace=True)
                    
                    cc_final_data = (
                        [df_cc_existing.columns.tolist()]
                        + df_cc_existing.astype(str).values.tolist()
                    )
                    cc_ws.clear()
                    cc_ws.update(cc_final_data)
                    print(f"cc_ho_data appended with {len(df_moved_new)} new rows (including remarks).")
                else:
                    print("All moved rows already exist in cc_ho_data — nothing appended.")

        # ── Vectorized Update Engine (merge-based, avoids index mismatch) ──
        ho_cols = df_ho_existing.columns.tolist()
        remarks_cols = {col for col in ho_cols if col.startswith("remark")}

        cols_to_update = [
            col for col in df_filtered.columns
            if col != "service_id"
            and col not in remarks_cols
            and col in df_ho_existing.columns
        ]
        print(f"Columns being updated: {cols_to_update}")

        matched_ids = set(df_ho_existing["service_id"]) & set(df_filtered["service_id"])
        updated_count = len(matched_ids)
        df_ho_new = df_filtered[~df_filtered["service_id"].isin(matched_ids)].reset_index(drop=True).copy()

        df_updates = (
            df_filtered[df_filtered["service_id"].isin(matched_ids)][["service_id"] + cols_to_update]
            .drop_duplicates(subset="service_id")
            .reset_index(drop=True)
        )

        df_ho_existing = df_ho_existing.merge(
            df_updates,
            on="service_id",
            how="left",
            suffixes=("", "_new")
        )

        for col in cols_to_update:
            new_col = col + "_new"
            if new_col in df_ho_existing.columns:
                mask = df_ho_existing[new_col].notna() & (df_ho_existing[new_col] != "")
                df_ho_existing.loc[mask, col] = df_ho_existing.loc[mask, new_col]
                df_ho_existing.drop(columns=[new_col], inplace=True)

        if not df_ho_new.empty:
            df_ho_new = df_ho_new.reindex(columns=ho_cols, fill_value="")
            df_ho_existing = pd.concat([df_ho_existing, df_ho_new], ignore_index=True)

        final_data = [df_ho_existing.columns.tolist()] + df_ho_existing.fillna("").astype(str).values.tolist()
        ho_ws.clear()
        ho_ws.update(final_data)

        show_popup(
            f"HO Raw Data updated! {updated_count} rows updated, "
            f"{len(df_ho_new)} new rows added, {len(df_moved)} rows moved to cc_ho_data.",
            type="success"
        )

    except Exception as e:
        print(f"Error in update_ho_raw_data function: {e}")
        show_popup(f"Error While Updating HO Data: {e}", type="error")


# To moving 15+ calls in ch data to ho data on current date
def checking_call_age_ch_data():
    try:
        spreadsheet = connect_gsheet()
        ch_ws = spreadsheet.worksheet("CH Raw Data")
        ch_data = ch_ws.get_all_values()
        if len(ch_data) <= 1:
            show_popup("CH Raw Data sheet is empty!", type="info")
            return

        df_ch = pd.DataFrame(ch_data[1:], columns=ch_data[0])
        df_ch.columns = df_ch.columns.str.lower().str.strip().str.replace(" ", "_")

        # ── Age calculation ──
        df_ch["call_date"] = pd.to_datetime(df_ch["call_date"]).dt.normalize()
        print("The call date in df ch is:", df_ch["call_date"])
        current_date       = pd.Timestamp.now().normalize()   # real date in production
        print("Real current date is:", current_date)
        # current_date           = "2026-06-12"
        current_date = pd.to_datetime(current_date)
        # current_date = pd.to_datetime(current_date).dt.normalize()
        print("current date is:", current_date)

        df_ch["age_at_todays"] = (current_date - df_ch["call_date"]).dt.days.fillna(0).astype(int)
        print("Age at todays as:", df_ch["age_at_todays"])
        # ── Update ageing display columns on ALL rows ──
        # These three columns are refreshed every run so CH Raw Data always
        # shows the latest age, not the age at the time the row was first added.
        #
        #   current_date      → today's date as a readable string  (DD-MM-YYYY)
        #   age_reg_days      → integer number of days since call_date
        #   age_from_call_reg → human-readable label  e.g. "18 days"
        #
        # If the columns don't exist yet they are created automatically.
        # df_ch["current_date"]       = current_date.strftime("%d-%m-%Y")
        df_ch["today_date"]       =  current_date.strftime("%Y-%m-%d")
        df_ch["age_reg_days"]       = df_ch["age_at_todays"]
        df_ch["age_from_call_reg"]  = df_ch["age_at_todays"].astype(str) + " days"

        # ── Write refreshed ageing back to CH Raw Data (before move logic) ──
        # Drop the internal temp column — we only keep the three display columns above.
        df_ch_to_write = df_ch.drop(columns=["age_at_todays"], errors="ignore")
        ch_refresh_data = [df_ch_to_write.columns.tolist()] + df_ch_to_write.fillna("").astype(str).values.tolist()
        ch_ws.clear()
        ch_ws.update(ch_refresh_data)
        print("CH Raw Data ageing columns refreshed.")

        # ── Filter rows to move (age == 16) ──
        # age_16_data = df_ch[df_ch["age_at_todays"] == 16].copy()
        age_16_data = df_ch[df_ch["age_at_todays"] >= 15].copy()
        print("Rows moving CH → HO Raw Data:", age_16_data.shape)

        if age_16_data.empty:
            show_popup("Ageing updated. No CH records with age = 16 found today.", type="info")
            return

        moved_ids = set(age_16_data["service_id"].astype(str).str.strip())

        # Drop internal temp + keep display columns intact for HO as well
        age_16_data = age_16_data.drop(columns=["age_at_todays"], errors="ignore")

        # ── Load / create HO Raw Data sheet ──
        ho_sheet_name = "HO Raw Data"
        try:
            ho_ws   = spreadsheet.worksheet(ho_sheet_name)
            ho_data = ho_ws.get_all_values()

            if ho_data and len(ho_data) > 1:
                df_ho = pd.DataFrame(ho_data[1:], columns=ho_data[0])
                df_ho.columns = df_ho.columns.str.lower().str.strip().str.replace(" ", "_")
                df_ho = df_ho.loc[:, ~df_ho.columns.duplicated()].reset_index(drop=True)
                df_ho["service_id"] = df_ho["service_id"].astype(str).str.strip()
            else:
                df_ho = pd.DataFrame()

        except gspread.WorksheetNotFound:
            ho_ws = spreadsheet.add_worksheet(ho_sheet_name, rows=5000, cols=30)
            df_ho = pd.DataFrame()

        age_16_data["service_id"] = age_16_data["service_id"].astype(str).str.strip()

        # ── Write to HO Raw Data ──
        if df_ho.empty:
            final_ho = [age_16_data.columns.tolist()] + age_16_data.fillna("").astype(str).values.tolist()
            ho_ws.update(final_ho)
            rows_added = len(age_16_data)
            skipped    = 0
        else:
            # Add any new columns from CH that don't exist in HO yet
            new_cols = [c for c in age_16_data.columns if c not in df_ho.columns.tolist()]
            if new_cols:
                print(f"New columns added to HO Raw Data: {new_cols}")
                for col in new_cols:
                    df_ho[col] = ""

            all_ho_cols = df_ho.columns.tolist()
            age_16_data = age_16_data.reindex(columns=all_ho_cols, fill_value="")

            existing_ho_ids = set(df_ho["service_id"])
            new_rows        = age_16_data[~age_16_data["service_id"].isin(existing_ho_ids)].copy()
            skipped         = len(age_16_data) - len(new_rows)
            rows_added      = len(new_rows)

            if skipped:
                print(f"{skipped} service_id(s) already in HO Raw Data — skipped.")

            if not new_rows.empty:
                df_ho    = pd.concat([df_ho, new_rows], ignore_index=True)
                final_ho = [df_ho.columns.tolist()] + df_ho.fillna("").astype(str).values.tolist()
                ho_ws.clear()
                ho_ws.update(final_ho)

        # ── Delete moved rows from CH Raw Data ──
        df_ch_remaining = df_ch_to_write[
            ~df_ch_to_write["service_id"].astype(str).str.strip().isin(moved_ids)
        ].copy()

        print(f"CH Raw Data: {len(df_ch_to_write)} rows before → {len(df_ch_remaining)} rows after removal.")

        ch_final = [df_ch_remaining.columns.tolist()] + df_ch_remaining.fillna("").astype(str).values.tolist()
        ch_ws.clear()
        ch_ws.update(ch_final)

        show_popup(
            f"Ageing refreshed. {rows_added} record(s) moved CH → HO Raw Data, "
            f"{len(moved_ids)} deleted from CH."
            + (f" ({skipped} duplicate(s) skipped in HO.)" if skipped else ""),
            type="success"
        )

    except Exception as e:
        print(f"Error in checking_call_age_ch_data: {e}")
        show_popup(f"Error in CH Age Check: {e}", type="error")

# To assign cco in HO data
def update_call_assignment_in_ho(assigned_df: pd.DataFrame) -> bool:
    try:
        # ── 1. Fetch current HO Data ──────────────────────────────────────
        spreadsheet = connect_gsheet()
        ho_ws = spreadsheet.worksheet("HO Raw Data")
        ho_data = ho_ws.get_all_values()

        if not ho_data or len(ho_data) <= 1:
            raise ValueError("HO Raw Data sheet is empty or contains no rows.")

        headers = [col.strip() for col in ho_data[0]]
        normalized_headers = [col.lower().replace(" ", "_") for col in headers]

        # Read into DataFrame using original headers to preserve state
        ho_df = pd.DataFrame(ho_data[1:], columns=headers)
        
        # FIX: Safeguard against duplicate columns in HO sheet
        ho_df = ho_df.loc[:, ~ho_df.columns.duplicated()]
        
        # ── 2. Normalize Incoming Excel Data ─────────────────────────────
        assigned_df.columns = assigned_df.columns.str.lower().str.strip().str.replace(" ", "_")
        
        # FIX: Safeguard against duplicate columns in uploaded file
        assigned_df = assigned_df.loc[:, ~assigned_df.columns.duplicated()]
        
        if "service_id" not in assigned_df.columns or "code" not in assigned_df.columns:
            print("Error: Uploaded file missing 'service_id' or 'code' columns.")
            return False

        # Helper function to strip '.0' from float strings safely
        def clean_code_string(val):
            val_str = str(val).strip()
            if val_str.endswith('.0'):
                return val_str[:-2]
            return val_str

        assigned_df["service_id"] = assigned_df["service_id"].astype(str).str.strip()
        
        # FIX: Clean the incoming codes to remove .0 before mapping
        assigned_df["code"] = assigned_df["code"].astype(str).apply(clean_code_string)
        
        assigned_map = assigned_df.dropna(subset=["service_id"]).set_index("service_id")["code"].to_dict()

        # Find column mappings using case/space-insensitive match
        service_id_orig_col = headers[normalized_headers.index("service_id")]

        # ── 3. Handle 'Code' Column Positioning & Index ──────────────────
        if "code" not in normalized_headers:
            # If 'code' column doesn't exist, we must add it to the Google Sheet structure
            if "15+_calls" in normalized_headers:
                target_col_idx = normalized_headers.index("15+_calls") + 2  # 1-based index + 1 right after
                header_title = "Code"
            else:
                target_col_idx = len(headers) + 1
                header_title = "Code"
            
            # Insert column in the Google Sheet structure
            ho_ws.insert_cols([[header_title]], col=target_col_idx)
            
            # Re-fetch headers and indices to sync up
            ho_data = ho_ws.get_all_values()
            headers = [col.strip() for col in ho_data[0]]
            normalized_headers = [col.lower().replace(" ", "_") for col in headers]
            ho_df = pd.DataFrame(ho_data[1:], columns=headers)
            ho_df = ho_df.loc[:, ~ho_df.columns.duplicated()] # Re-apply deduplication

        # Get exact 1-based index of the Code column
        code_col_idx = normalized_headers.index("code") + 1
        code_orig_col = headers[code_col_idx - 1]

        # ── 4. Match and Update local DataFrame column ───────────────────
        ho_df[service_id_orig_col] = ho_df[service_id_orig_col].astype(str).str.strip()
        
        # Calculate new codes while safely handling preexisting values
        ho_df[code_orig_col] = ho_df[service_id_orig_col].map(assigned_map).fillna(ho_df[code_orig_col]).replace("nan", "")
        
        # FIX: Also clean pre-existing codes in the sheet just in case they contain '.0'
        ho_df[code_orig_col] = ho_df[code_orig_col].astype(str).apply(clean_code_string).replace("nan", "")

        # ── 5. Push ONLY the Code Column back to Google Sheets ───────────
        # Extract only the calculated Code values as a list of lists (column format)
        code_values_to_upload = [[val] for val in ho_df[code_orig_col].tolist()]
        
        # Define range starting from Row 2 (skipping header) to the last row
        start_row = 2
        end_row = start_row + len(code_values_to_upload) - 1
        
        # Format update range dynamically using column indexes (e.g., "E2:E100")
        range_start = rowcol_to_a1(start_row, code_col_idx)
        range_end = rowcol_to_a1(end_row, code_col_idx)
        update_range = f"{range_start}:{range_end}"

        # Single batch update targeted strictly at that range
        ho_ws.update(range_name=update_range, values=code_values_to_upload)
        
        show_popup("Successfully updated", type = "success")
        return True
        
    except Exception as e:
        print(f"Backend processing error: {e}")
        show_popup(f"Backend processing error: {e}", type= "error")
        return False



# To create summmary report
def calls_data():
    try:
        spreadsheet = connect_gsheet()
        detail_ws = spreadsheet.worksheet("Detailed_Data")
        detail_data = detail_ws.get_all_values()
        if len(detail_data) <= 1:
            show_popup("Detailed_Data sheet is empty!", type="info")
            return None

        df_detail = pd.DataFrame(detail_data[1:], columns=detail_data[0])
        df_detail.columns = df_detail.columns.str.lower().str.strip().str.replace(" ", "_")
        df_detail = df_detail.loc[:, ~df_detail.columns.duplicated()]
        
        # 1. Convert call columns to numeric safely
        df_detail['7+_calls'] = pd.to_numeric(df_detail['7+_calls'], errors='coerce').fillna(0)
        df_detail['15+_calls'] = pd.to_numeric(df_detail['15+_calls'], errors='coerce').fillna(0)
        df_detail["circle"] = df_detail["circle"].str.strip().str.lower()
        df_detail["status_code"] = df_detail["status_code"].str.strip().str.upper()

        # 2. Groupby 'circle' and 'status_code' and aggregate
        df_summary = df_detail.groupby(['circle', 'status_code']).agg(
            total_7plus_count=('7+_calls', lambda x: (x > 0).sum()),
            total_14plus_count=('15+_calls', lambda x: (x > 0).sum())
        ).reset_index()

        # 3. Add the Combined Total Column
        df_summary['grand_total_calls'] = df_summary['total_7plus_count'] + df_summary['total_14plus_count']
        df_summary["circle"] = df_summary["circle"].str.title()

        # ---- ADDING THE TOTAL ROW AT THE BOTTOM ----
        if not df_summary.empty:
            # Create a dictionary representing the total row layout
            total_row = pd.DataFrame([{
                'circle': 'Total',
                'status_code': '',  # Kept blank for aesthetic neatness
                'total_7plus_count': df_summary['total_7plus_count'].sum(),
                'total_14plus_count': df_summary['total_14plus_count'].sum(),
                'grand_total_calls': df_summary['grand_total_calls'].sum()
            }])
            
            # Combine the summary dataframe with the new total row safely
            df_summary = pd.concat([df_summary, total_row], ignore_index=True)

        return df_summary

    except Exception as e:
        print(f"Error in calls data function is: {e}")
        show_popup(f"Error in calls data function is: {e}")
        return None


# --- Excel Formatting Helper Function ---
def convert_df_to_formatted_excel(df):
    output = io.BytesIO()
    
    # Write DataFrame to excel
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Summary_Report')
        
        # Grab the worksheet to apply styling
        workbook = writer.book
        worksheet = writer.sheets['Summary_Report']
        
        # Define some clean, professional styles
        header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid') # Deep steel blue
        data_font = Font(name='Segoe UI', size=10)
        
        # Alignment & Borders
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )
        
        # Format Headers
        for col_num in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=1, column=col_num)
            # Make column names look clean (e.g. "total_7plus_count" -> "Total 7Plus Count")
            cell.value = str(cell.value).replace('_', ' ').title()
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
            
        # Format Data Rows & Columns
        for row in range(2, worksheet.max_row + 1):
            for col in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.font = data_font
                cell.border = thin_border
                
                # Align numeric data to center, text/codes to left
                if col in [1, 2]:  # Circle & Status Code
                    cell.alignment = left_align
                else:              # Counts
                    cell.alignment = center_align
                    cell.number_format = '#,##0'  # Clean integer formatting

        # Auto-adjust column widths dynamically based on content length
        for col in worksheet.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            worksheet.column_dimensions[col_letter].width = max(max_len + 4, 12)

    processed_data = output.getvalue()
    return processed_data
